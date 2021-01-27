from csv import DictWriter
from openpyxl import Workbook
from datetime import date
from io import BytesIO

from .forms import RegisterUserForm, UserLoginForm, UserProfileForm, UserProfileFormActive, IndexTaskForm, TaskForm, HolidayForm, CommentForm
from .models import Task, Engineer, Holiday, Comment, Following, Follower
from .filters import TaskFilter
from .utils import render_to_pdf

from django.shortcuts import render, redirect, reverse
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.views.generic import View, TemplateView
# from django.db.models import Count
from xhtml2pdf import pisa


@login_required(login_url="login")
def index(request):

    if not request.user.is_staff:
        engineer = request.user.engineer
        task_counter = Task.objects.filter(assigned_to=engineer).count()
        pending_task = Task.incompleted_task.filter(assigned_to=engineer).count()
        completed_task = Task.completed_task.filter(assigned_to=engineer).count()
        all_task = Task.objects.filter(assigned_to__exact=engineer)

        form = IndexTaskForm(initial={"assigned_to":engineer})
    else:
        all_task = Task.objects.all()
        task_counter = all_task.count()
        pending_task = Task.incompleted_task.all().count()
        completed_task = Task.completed_task.all().count()

        form = IndexTaskForm()

    myfilter = TaskFilter(request.GET, queryset=all_task)
    all_task = myfilter.qs

    paginator = Paginator(all_task, 5)
    page_num = request.GET.get("page")
    page_obj = paginator.get_page(page_num)
    
    if request.method == "POST":
        form = IndexTaskForm(request.POST)
        if form.is_valid():
            return redirect("create_task")

    context = {
        "task_counter":task_counter,
        "pending_task":pending_task,
        "completed_task":completed_task,
        "all_task":all_task,
        "form":form,
        "myfilter":myfilter,
        "page_obj":page_obj,
        }

    return render(request, "task_app/index.html", context)

@login_required(login_url="login")
def user_search(request):
    if request.method == "POST":
        search_for = request.POST.get("search")
        engineers = Engineer.objects.filter(owner__username__icontains=search_for)

        context = {"engineers":engineers}

        return render(request,"task_app/engineer_list.html", context)

@login_required(login_url="login")
def engineer_profile(request, username):
    engineer = Engineer.objects.get(owner__username=username)
    tasks = Task.objects.filter(assigned_to=engineer)
    follower = engineer.follower.all().count()
    following = engineer.following.all().count()

    is_followed = Follower.objects.filter(owner_id=engineer, follower_id=request.user.engineer)

    context = {"engineer":engineer, "tasks":tasks, "following":following, "follower":follower, "is_followed":is_followed}
    return render(request, "task_app/user_info.html", context)

@login_required(login_url="login")
def follow_profile(request, username):
    engineer = request.user.engineer
    engineer_to_follow = Engineer.objects.get(owner__username=username)

    Following.objects.create(owner_id=engineer, following_id=engineer_to_follow)
    Follower.objects.create(owner_id=engineer_to_follow, follower_id=engineer)

    return redirect(reverse("user_info", kwargs={"username":username}))

@login_required(login_url="login")
def unfollow_profile(request, username):
    engineer = request.user.engineer
    engineer_to_follow = Engineer.objects.get(owner__username=username)

    Following.objects.filter(owner_id=engineer, following_id=engineer_to_follow).delete()
    Follower.objects.filter(owner_id=engineer_to_follow, follower_id=engineer).delete()

    return redirect(reverse("user_info", kwargs={"username":username}))

@login_required(login_url="login")
def task_comments(request, pk):
    if request.method == "POST":
        form = CommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.owner = request.user
            comment.task = Task.objects.get(id=pk)
            comment.save()
    else:
        form = CommentForm()

    comments = Comment.objects.filter(task=pk)
    actual_task = Task.objects.get(id=pk)

    context = {"comments":comments, "form":form, "task":actual_task}
    return render(request, "task_app/task_comments.html", context)

@login_required(login_url="login")
def delete_comment(request, pk):
    comment = Comment.objects.get(id=pk)
    # for redirection to task_comments page
    task_id = comment.task.id
    comment.delete()
    #return redirect(reverse("task_comments", kwargs={"pk":task_id}))
    return redirect("task_comments", pk=task_id)
    #return redirect("task_comments")

@login_required(login_url="login")
def update_comment(request, pk):
    comment = Comment.objects.get(id=pk)
    # for redirection to task_comments page
    task_id = comment.task.id
    if request.method == "POST":
        form = CommentForm(request.POST, instance=comment)
        if form.is_valid():
            form.save()
            return redirect(reverse("task_comments", kwargs={"pk":task_id}))
    else:
        form = CommentForm(instance=comment)
    
    context = {"form":form}
    return render(request, "task_app/task_comments_update.html", context)

@login_required(login_url="login")
def export_csv(request):
    response = HttpResponse(content_type="text/csv")

    fieldnames = ["Assigned To", "Category", "Date"]
    writer = DictWriter(response, fieldnames=fieldnames)

    tasks = Task.objects.filter(assigned_to=request.user.engineer).values_list("assigned_to", "category", "start_date")

    writer.writeheader()
    for task in tasks:
        writer.writerow({"Assigned To":Engineer.objects.get(id=task[0]), "Category":task[1], "Date":task[2]})
    
    response["Content-Disposition"] = 'attachment; filename="tasks.csv"'

    return response

@login_required(login_url="login")
def export_excel(request):
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="tasks.xlsx"'

    col_names = ["Assigned To", "Category", "Date"]
    tasks = Task.objects.filter(assigned_to=request.user.engineer).values_list("assigned_to", "category", "start_date")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Tasks of {request.user.engineer}"
    
    for col in range(1,len(col_names)+1):
        ws.cell(row=1, column=col, value=col_names[col-1])

    last_row = ws.max_row + 1
    for row in range(last_row,len(tasks)+last_row):
        for col in range(1,len(col_names)+1):
            ws.cell(row=row,column=col, value=tasks[row-last_row][col-1])
    
    wb.save(response)
    return response


class GeneratePDF(View):
    def get(self, request, *args, **kwargs):
        task = Task.objects.get(id=37)
        comments = Comment.objects.filter(task=37)

        context = {
            "task": task,
            "comments": comments,
        }

        pdf = render_to_pdf('task_app/task_comments.html', context)
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "Invoice_37.pdf"
            content = f"inline; filename={filename}"
            download = request.GET.get("download")
            if download:
                content = f"attachment; filename={filename}"
            response['Content-Disposition'] = content
            return response
        return HttpResponse("Not found")

class StatisticsView(TemplateView):
    template_name = 'task_app/statistics.html'

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        labels = [category[0] for category in Task.categories]
        context["category_labels"] = labels

        tasks = Task.objects.filter(assigned_to=self.request.user.engineer)
        data = [tasks.filter(category=label).count() for label in labels]
        context["data"] = data

        return context

def register_user(request):
    if request.method == "POST":
        form = RegisterUserForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "New Profile has been created.")
            return redirect("login")
        else:
            messages.info(request, "Something went wrong during registration. Try again.")
    else:
        form = RegisterUserForm()

    context = {"form":form}
    return render(request, "registration/register.html", context)

def login_user(request):
    if request.method == "POST":
        form = UserLoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data["username"]
            password = form.cleaned_data["password"]

            user = authenticate(request, username=username, password=password)

            if user is not None:
                login(request, user)
                messages.success(request, "Welcome. You have logged in successfully.")
                return redirect("index")
            else:
                messages.info(request, "Username or Password incorrect. Try again.")
        else:
             messages.info(request, "Username or Password incorrect. Try again.")
    else:
        form = UserLoginForm()

    context = {"form":form}
    return render(request, "registration/login.html", context)

@login_required(login_url="login")
def logout_user(request):
    logout(request)
    messages.success(request, "You have logged out successfully.")
    return redirect("login")

@login_required(login_url="login")
def user_profile(request):
    engineer = request.user.engineer
    form = UserProfileForm(instance=engineer)

    context = {"form":form}
    return render(request,"task_app/profile.html", context)

@login_required(login_url="login")
def user_profile_update(request):
    engineer = request.user.engineer
    if request.method == "POST":
        form = UserProfileFormActive(request.POST, request.FILES, instance=engineer)
        if form.is_valid():
            form.save()
            return redirect("profile")
    else:
        form = UserProfileFormActive(instance=engineer)

    context = {"form":form}
    return render(request,"task_app/profile.html", context)

@login_required(login_url="login")
def user_holidays(request):   
    if not request.user.is_staff:
        engineer = request.user.engineer
        holidays = Holiday.objects.filter(owner=engineer)
        if request.method == "POST":
            form = HolidayForm(request.POST, initial={"owner":engineer})
            if form.is_valid():
                form.save()
                return redirect("holidays")
        else:
            form = HolidayForm(initial={"owner":engineer})
            
    else:
        holidays = Holiday.objects.all()
        if request.method == "POST":
            form = HolidayForm(request.POST)
            if form.is_valid():
                form.save()
                return redirect("holidays")  
        else:
            form = HolidayForm()      

    context = {"form":form, "holidays":holidays}
    return render(request,"task_app/holidays.html", context)

@login_required(login_url="login")
def update_holidays(request, pk):
    holiday = Holiday.objects.get(id=pk)
    if request.method == "POST":
        form = HolidayForm(request.POST, instance=holiday)
        if form.is_valid():
            form.save()
            return redirect("holidays")
    else:
        form = HolidayForm(instance=holiday)

    context = {"form":form}
    return render(request,"task_app/update_holidays.html", context)

@login_required(login_url="login")
def delete_holidays(request, pk):
    holiday = Holiday.objects.get(id=pk)
    holiday.delete()
    return redirect("holidays")

@login_required(login_url="login")
def task_creation(request):
    engineer = request.POST.get("assigned_to")
    start_date = request.POST.get("start_date")
    # engineer = request.user.engineer
    # start_date = date.today()

    if request.method == "POST":
        form = TaskForm(request.POST)
        if form.is_valid():
            form.save()
            messages.info(request, "New task has been created.")
            return redirect("index")
    else:   
        form = TaskForm(initial={"assigned_to":engineer, "start_date":start_date})

    context = {"form":form}

    return render(request,"task_app/create_new_task.html", context)

@login_required(login_url="login")
def update_task(request, pk):
    task = Task.objects.get(id=pk)

    if request.method == "POST":
        form = TaskForm(request.POST, instance=task)
        if form.is_valid():
            form.save()
            messages.info(request, "Task has been updated.")
            return redirect("index")
    else:       
        form = TaskForm(instance=task)

    context = {"form":form}

    return render(request,"task_app/update_task.html", context)

@login_required(login_url="login")
def delete_task(request, pk):
    task = Task.objects.get(id=pk)
    task.delete()
    return redirect("index")
